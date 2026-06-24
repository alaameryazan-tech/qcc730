
#for xls
import xlrd
import xlwt

#for xlsx write
import xlsxwriter

#for xlsx read/write, recommended
import openpyxl
import os
import re

NAME_PERIOD = ord('Z') - ord('A') + 1

#num starts from 0
#return 0~25
def col_chr2num(char):
    if ((ord(char)<ord('A')) or (ord(char)>ord('Z'))):
        print('char=%c is not supported'%char)
        exit(1)
    return ord(char) - ord('A')

#num starts from 0
#return 'A'~'Z'
def col_num2chr(num):
    if ((num<0) or (num>NAME_PERIOD-1)):
        print('num=%d is not supported'%num)
    return chr(ord('A')+num)

#only support 2 chars
#num starts from 0
#retrun 0~(26*26+25)
def col_name2num(colname):
    char_count = len(colname)
    if char_count==1:
        return col_chr2num(colname)
    elif char_count==2:
        return (NAME_PERIOD*(col_chr2num(colname[0])+1) + col_num2chr(colname[1]))
    else:
        print('colname=%s is too long, not supported'%colname)
        exit(1)

#only support 2 chars
#num starts from 0
#return 'A'~'ZZ'
def col_num2name(colnum):
    if colnum < NAME_PERIOD:
        return col_num2chr(colnum)
    elif colnum < NAME_PERIOD*NAME_PERIOD:
        ch1 = col_num2chr(int(colnum/NAME_PERIOD)-1)
        ch2 = col_num2chr(colnum%NAME_PERIOD)
        return ch1+ch2
    else:
        print('colnum=%d is too large, not supported'%colnum)
        exit(1)

#openpyxl row starts from 1
#expect user row starts from 0
def orow(n):
    return (n+1)

#openpyxl col starts from 1
#expect user col starts from 0
def ocol(n):
    return (n+1)

#only support 2 chars for col, so max col is 'ZZ'
#col/row num starts from 0
#return 'Ax'~ZZx'', 'x' starts from '1'
def cellname(colnum, rownum):
    return col_num2name(colnum)+str(rownum+1)

#col_name starts from 'A', row_name starts from '1'
def cellname_split(name):
    found = re.search('([A-Z]+)', name)
    col_name = found.group(1)
    found = re.search('([0-9]+)', name)
    row_name = found.group(1)
    return (col_name, row_name)

#for original xlsx read/write on openpyxl, rol/col starts from 1; here transform to start from 0
class openxl (object):
    def __init__ (self, path, sheet_name):
        self.path = path
        self.sheet_name = sheet_name
        self.rows_count = 0
        self.cols_count = 0
        if os.path.isfile(path):
            self.book = openpyxl.load_workbook(path)
            self.sheet = self.book.get_sheet_by_name(sheet_name)
            if self.sheet==None:
                #create a new sheet
                self.sheet = self.book.create_sheet(title=sheet_name)
                self.book.save(self.path)
            else:
                #both book and sheet exist
                self.rows_count = self.sheet.max_row
                self.cols_count = self.sheet.max_column
        else:
            #create a new book
            self.book = openpyxl.Workbook()
            self.sheet = self.book.active
            self.sheet.title = sheet_name
            self.book.save(self.path)
        self.font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE, bold=True)

    def read_cell(self, rowid, colid):
        return self.sheet.cell(orow(rowid), ocol(colid)).value

    def read_row(self, rowid):
        row_values = []
        for row in self.sheet.iter_rows(min_col=1, min_row=orow(rowid), max_col=self.cols_count, max_row=orow(rowid)):
            for cell in row:
                row_values.append(cell.value)
        return row_values

    def read_col(self, colid):
        col_values = []
        for col in self.sheet.iter_cols(min_col=ocol(colid), min_row=1, max_col=ocol(colid), max_row=self.rows_count):
            for cell in col:
                col_values.append(cell.value)
        return col_values

    def save(self):
        self.book.save(self.path)

    def write_cell(self, row, col, value):
        self.sheet.cell(orow(row), ocol(col), value)

    def write_title(self, row, col, value, width=15, height=20):
        self.sheet.cell(orow(row), ocol(col), value)
        self.sheet.cell(orow(row), ocol(col)).font = self.font
        self.sheet.column_dimensions[col_num2name(col)].width = width
        self.sheet.row_dimensions[orow(row)].height = height

    #donot know how width/height are measured yet, just try
    def linechart_create(self, title, xaxis_name, yaxis_name, x_min_col, x_min_row, x_max_col, x_max_row, width=None, height=None):
        self.linechart = openpyxl.chart.LineChart()
        self.linechart.title = title
        self.linechart.x_axis.title = xaxis_name
        self.linechart.y_axis.title = yaxis_name

        if width!=None:
            #self.linechart.width = self.sheet.column_dimensions[col_num2name(0)].width*6
            self.linechart.width = width    #cm
        if height!=None:
            #self.linechart.height = self.sheet.row_dimensions[orow(0)].height*25
            self.linechart.height = height  #cm
        self.x_refObj = openpyxl.chart.Reference(self.sheet, min_col=ocol(x_min_col), min_row=orow(x_min_row), max_col=ocol(x_max_col), max_row=orow(x_max_row))

    #titles_from_data = True
    def linechart_add_ydata(self, min_col, min_row, max_col, max_row, title_str=None):
        refObj = openpyxl.chart.Reference(self.sheet, min_col=ocol(min_col), min_row=orow(min_row), max_col=ocol(max_col), max_row=orow(max_row))
        if title_str==None:
            self.linechart.add_data(refObj, titles_from_data=True)
        else:
            series_obj = openpyxl.chart.Series(refObj, title=title_str)
            self.linechart.append(series_obj)

    def linechart_save(self, row, col):
        self.linechart.set_categories(self.x_refObj)
        self.sheet.add_chart(self.linechart, cellname(col, row))

#for xls write, row/col starts from 0
#but row name start from '1', col name starts from 'A'
class xlsxwriter_write(object):
    def __init__ (self, path, sheet_name):
        if os.path.isfile(path):
            os.remove(path)
        self.path = path
        self.sheet_name = sheet_name
        self.book = xlsxwriter.Workbook(self.path)
        self.sheet = self.book.add_worksheet(sheet_name)
        self.style_bold = self.book.add_format({'bold': True})
        self.linechart = None
        self.x_start_name = None
        self.x_end_name = None

    def save(self):
        self.book.close()

    def write_cell(self, row, col, value):
        self.sheet.write(row, col, value)

    def write_title(self, row, col, value, width=15):
        self.sheet.write(row, col, value, self.style_bold)
        self.sheet.set_column(col, col , width)

    #width/height is multiple of cell[0][0]
    def linechart_create(self, title, xaxis_name, yaxis_name, xaxis_start_cell_name, xaxis_end_cell_name, width=None, height=None):
        self.linechart = self.book.add_chart({'type': 'line'})
        if width!=None:
            self.linechart.set_size({'width': width*self.sheet._size_col(0)})
        if height!=None:
            self.linechart.set_size({'height': (height*self.sheet._size_row(0))})
        self.linechart.set_title({'name': title})
        self.linechart.set_x_axis({'name': xaxis_name})
        self.linechart.set_y_axis({'name': yaxis_name})
        self.x_start_name = xaxis_start_cell_name
        (self.x_start_colname, self.x_start_rowname) =  cellname_split(xaxis_start_cell_name)
        self.x_end_name = xaxis_end_cell_name
        (self.x_end_colname, self.x_end_rowname) =  cellname_split(xaxis_end_cell_name)

    def linechart_add_ydata(self, title_cell_name, start_cell_name, end_cell_name):
        (title_colname, title_rowname) = cellname_split(title_cell_name)
        (start_colname, start_rowname) = cellname_split(start_cell_name)
        (end_colname, end_rowname) = cellname_split(end_cell_name)
        self.linechart.add_series({
                'name': '=%s!$%s$%s'%(self.sheet_name, title_colname, title_rowname),
                'categories': '=%s!$%s$%s:$%s$%s'%(self.sheet_name, self.x_start_colname, self.x_start_rowname, self.x_end_colname, self.x_end_rowname),
                'values':   '=%s!$%s$%s:$%s$%s'%(self.sheet_name, start_colname, start_rowname, end_colname, end_rowname)
            })

    def linechart_save(self, row, col):
        self.sheet.insert_chart(cellname(col, row), self.linechart)

#for xls write, row/col starts from 0
class xlwt_write(object):
    def __init__ (self, path, sheet_name):
        self.style = self.format_style()
        if os.path.isfile(path):
            os.remove(path)
        self.path = path
        self.book = xlwt.Workbook()
        self.sheet = self.book.add_sheet(sheet_name)
        self.book.save(self.path)

    def format_style(self):
        style = xlwt.XFStyle()
        font = xlwt.Font()
        font.name = ''
        font.bold = True
        font.colour_index = 4
        font.height = 200
        style.font = font
        return style

    def save(self):
        self.book.save(self.path)

    def write_cell(self, row, col, value):
        self.sheet.write(row, col, value)

    def write_title(self, row, col, value, width=256*15):
        self.sheet.write(row, col, value, self.style)
        self.sheet.col(col).width = width

#for xls read, row/col starts from 0
class xlrd_read(object):
    def __init__ (self, path, sheet_name):
        self.book = xlrd.open_workbook(path)
        self.sheet = self.book.sheet_by_name(sheet_name)
        self.rows_count = self.sheet.nrows
        self.cols_count = self.sheet.ncols

    def read_cell(self, rowid, colid):
        return self.sheet.cell_value(rowid, colid)

    def read_row(self, rowid):
        return self.sheet.row_values(rowid, start_colx=0, end_colx=None)

    def read_col(self, colid):
        return self.sheet.col_values(colid, start_rowx=0, end_rowx=None)

